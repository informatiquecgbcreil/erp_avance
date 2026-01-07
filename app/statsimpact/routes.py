from __future__ import annotations

from datetime import date
from io import BytesIO

import os

from flask import Blueprint, abort, render_template, request, redirect, url_for, flash, current_app, send_file
from flask_login import login_required, current_user

from sqlalchemy import func

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.extensions import db

from app.models import AtelierActivite, Participant, Quartier, PresenceActivite, SessionActivite

from .occupancy import compute_occupancy_stats

from .engine import (
    compute_volume_activity_stats,
    compute_participation_frequency_stats,
    compute_transversalite_stats,
    compute_demography_stats,
    compute_participants_stats,
    compute_magatomatique,
    normalize_filters,
)

bp = Blueprint("statsimpact", __name__, url_prefix="")


def _can_view() -> bool:
    return getattr(current_user, "role", None) in (
        "finance",
        "financiere",
        "financière",
        "directrice",
        "responsable_secteur",
        "admin_tech",
    )


def _safe_sheet_title(name: str, fallback: str = "Atelier") -> str:
    """Openpyxl: max 31 chars, no [ ] * ? / \\ etc."""
    if not name:
        name = fallback
    bad = set('[]:*?/\\')
    cleaned = "".join(c for c in name if c not in bad).strip()
    cleaned = cleaned[:31] if cleaned else fallback
    return cleaned


def _build_magato_per_atelier_workbook(flt) -> Workbook:
    """Export annuel type "Excel historique" : 1 feuille par atelier (matrice participants x sessions)."""

    # Cloisonnement : un responsable_secteur ne doit exporter que son secteur
    role = getattr(current_user, "role", None)
    eff_secteur = flt.secteur
    if role == "responsable_secteur":
        eff_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip() or eff_secteur

    # Liste des ateliers dans le périmètre
    aq = AtelierActivite.query.filter(AtelierActivite.is_deleted.is_(False))
    if eff_secteur:
        aq = aq.filter(AtelierActivite.secteur == eff_secteur)
    ateliers = aq.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()

    wb = Workbook()
    # on réutilise la 1ère feuille pour une synthèse
    ws0 = wb.active
    ws0.title = "Synthese"
    ws0.append(["Export annuel : 1 feuille par atelier"])
    ws0.append(["Secteur", "Atelier", "Nb sessions", "Nb présences", "Participants uniques", "Nouveaux", "Récurrents"])

    for at in ateliers:
        # Sessions de l'atelier dans la période
        sess_q = (
            db.session.query(SessionActivite)
            .filter(SessionActivite.atelier_id == at.id)
        )
        # filtre dates (inclusif)
        if flt.date_from:
            sess_q = sess_q.filter(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session) >= flt.date_from)
        if flt.date_to:
            sess_q = sess_q.filter(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session) <= flt.date_to)
        sess_q = sess_q.order_by(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session).asc(), SessionActivite.id.asc())
        sessions = sess_q.all()
        if not sessions:
            # atelier sans sessions dans la période -> on le garde dans la synthèse avec 0
            ws0.append([at.secteur, at.nom, 0, 0, 0, 0, 0])
            continue

        session_ids = [s.id for s in sessions]

        # Presences (pairs pid/sid)
        pres_rows = (
            db.session.query(PresenceActivite.participant_id, PresenceActivite.session_id)
            .filter(PresenceActivite.session_id.in_(session_ids))
            .all()
        )
        if not pres_rows:
            ws0.append([at.secteur, at.nom, len(sessions), 0, 0, 0, 0])
            # feuille vide mais structurée
            ws = wb.create_sheet(_safe_sheet_title(f"{at.nom}"))
            ws.append([f"{at.secteur} — {at.nom}"])
            ws.append(["Nom", "Prénom"] + [( ( (s.rdv_date or s.date_session).strftime("%d/%m/%Y") ) if (s.rdv_date or s.date_session) else "Sans date") for s in sessions])
            continue

        pid_set = sorted({int(pid) for (pid, _) in pres_rows if pid is not None})

        # Participants (id, nom, prénom)
        parts = (
            db.session.query(Participant.id, Participant.nom, Participant.prenom)
            .filter(Participant.id.in_(pid_set))
            .order_by(Participant.nom.asc(), Participant.prenom.asc())
            .all()
        )

        # Comptes / min date par participant pour KPI nouveaux/récurrents
        counts = (
            db.session.query(
                PresenceActivite.participant_id.label("pid"),
                func.count(PresenceActivite.id).label("nb"),
                func.min(func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session)).label("first"),
            )
            .select_from(PresenceActivite)
            .join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
            .filter(PresenceActivite.session_id.in_(session_ids))
            .group_by(PresenceActivite.participant_id)
            .all()
        )
        c_map = {int(r.pid): {"nb": int(r.nb or 0), "first": r.first} for r in counts if r and r.pid is not None}
        new_count = 0
        recurring = 0
        for pid in pid_set:
            nb = int(c_map.get(pid, {}).get("nb", 0))
            if nb >= 2:
                recurring += 1
            fd = c_map.get(pid, {}).get("first")
            if fd and flt.date_from and flt.date_to and flt.date_from <= fd <= flt.date_to:
                new_count += 1

        ws0.append([at.secteur, at.nom, len(sessions), len(pres_rows), len(pid_set), new_count, recurring])

        # Matrice
        ws = wb.create_sheet(_safe_sheet_title(f"{at.nom}"))
        ws.append([f"{at.secteur} — {at.nom}"])
        headers = ["Nom", "Prénom"] + [
            ((d.strftime("%d/%m/%Y")) if (d := (s.rdv_date or s.date_session)) else "Sans date")
            for s in sessions
        ]
        ws.append(headers)

        # index session -> col offset
        sid_index = {int(s.id): idx for idx, s in enumerate(sessions)}
        present = set((int(pid), int(sid)) for (pid, sid) in pres_rows if pid is not None and sid is not None)

        for pid, nom, prenom in parts:
            row = [nom or "", prenom or ""] + [""] * len(sessions)
            for sid, idx in sid_index.items():
                if (int(pid), int(sid)) in present:
                    row[2 + idx] = "1"
            ws.append(row)

        # Largeurs raisonnables
        ws.column_dimensions[get_column_letter(1)].width = 20
        ws.column_dimensions[get_column_letter(2)].width = 18
        for col_idx in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

    return wb


@bp.route("/stats-impact", methods=["GET", "POST"])
@login_required
def dashboard():
    if not _can_view():
        abort(403)

    args = dict(request.args)

    # Robust: normalize_filters supports both dict-style and kwargs-style.
    flt = normalize_filters(args, user=current_user)

    # Default: current year if no dates
    if not flt.date_from and not flt.date_to:
        today = date.today()
        flt.date_from = date(today.year, 1, 1)
        flt.date_to = date(today.year, 12, 31)

    # Pre-compute participants for access control if we need to handle edits.
    participants = compute_participants_stats(flt)

    if request.method == "POST":
        action = request.form.get("action")
        if action == "update_participant":
            try:
                participant_id = int(request.form.get("participant_id", "0"))
            except Exception:
                participant_id = 0

            allowed_ids = {p["id"] for p in participants.get("participants", [])}
            if not participant_id or participant_id not in allowed_ids:
                abort(403)

            participant = Participant.query.get(participant_id)
            if not participant:
                abort(404)

            participant.nom = (request.form.get("nom") or participant.nom or "").strip() or participant.nom
            participant.prenom = (request.form.get("prenom") or participant.prenom or "").strip() or participant.prenom
            participant.ville = (request.form.get("ville") or "").strip() or None
            participant.email = (request.form.get("email") or "").strip() or None
            participant.telephone = (request.form.get("telephone") or "").strip() or None
            participant.genre = (request.form.get("genre") or "").strip() or None
            participant.type_public = (request.form.get("type_public") or participant.type_public or "H").strip().upper()

            dn_raw = request.form.get("date_naissance") or None
            dn = None
            if dn_raw:
                try:
                    dn = date.fromisoformat(dn_raw)
                except Exception:
                    dn = None
            participant.date_naissance = dn

            quartier_id = request.form.get("quartier_id") or None
            try:
                participant.quartier_id = int(quartier_id) if quartier_id else None
            except Exception:
                participant.quartier_id = None

            try:
                from app.extensions import db

                db.session.commit()
                flash("Participant mis à jour.", "success")
            except Exception:
                db.session.rollback()
                flash("Impossible de sauvegarder ce participant.", "danger")

            args_redirect = request.args.to_dict(flat=True)
            args_redirect["tab"] = "participants"
            return redirect(url_for("statsimpact.dashboard", **args_redirect))

        if action == "delete_participant":
            try:
                participant_id = int(request.form.get("participant_id", "0"))
            except Exception:
                participant_id = 0

            allowed_ids = {p["id"] for p in participants.get("participants", [])}
            if not participant_id or participant_id not in allowed_ids:
                abort(403)

            participant = Participant.query.get(participant_id)
            if not participant:
                abort(404)

            # Sécurité secteur: un responsable_secteur ne peut purger un participant
            # que si ce participant n'a des présences que dans SON secteur (ou aucune).
            role = getattr(current_user, "role", None)
            user_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip()
            if role == "responsable_secteur":
                sectors = (
                    PresenceActivite.query.join(SessionActivite, PresenceActivite.session_id == SessionActivite.id)
                    .with_entities(SessionActivite.secteur)
                    .filter(PresenceActivite.participant_id == participant_id)
                    .distinct()
                    .all()
                )
                sectors = {s[0] for s in sectors if s and s[0]}
                # s'il n'a jamais émargé: OK (secteurs = vide)
                if sectors and sectors != {user_secteur}:
                    flash(
                        "Suppression refusée : ce participant a des émargements dans d'autres secteurs.",
                        "danger",
                    )
                    args_redirect = request.args.to_dict(flat=True)
                    args_redirect["tab"] = "participants"
                    return redirect(url_for("statsimpact.dashboard", **args_redirect))

            try:
                from app.extensions import db

                # Supprime d'abord les signatures des présences
                presences = PresenceActivite.query.filter_by(participant_id=participant_id).all()
                for pr in presences:
                    if pr.signature_path:
                        try:
                            if os.path.exists(pr.signature_path):
                                os.remove(pr.signature_path)
                        except Exception:
                            pass
                    db.session.delete(pr)

                db.session.delete(participant)
                db.session.commit()
                flash("Participant supprimé définitivement.", "success")
            except Exception:
                db.session.rollback()
                flash("Impossible de supprimer ce participant.", "danger")

            args_redirect = request.args.to_dict(flat=True)
            args_redirect["tab"] = "participants"
            return redirect(url_for("statsimpact.dashboard", **args_redirect))

    # Refresh computed stats after any potential mutation
    participants = compute_participants_stats(flt)
    stats = compute_volume_activity_stats(flt)
    freq = compute_participation_frequency_stats(flt)
    trans = compute_transversalite_stats(flt)
    demo = compute_demography_stats(flt)
    occupancy = compute_occupancy_stats(flt)

    # Le Magatomatique : calcul uniquement si l'onglet est affiché (sinon on garde la page légère)
    tab = (request.args.get("tab") or "base").strip().lower()
    magato = None
    if tab in ("magato", "magatomatique"):
        participant_q = (request.args.get("participant_q") or "").strip() or None
        view = (request.args.get("magato_view") or "macro").strip().lower()
        try:
            max_sessions = int(request.args.get("max_sessions") or 40)
        except Exception:
            max_sessions = 40
        try:
            max_participants = int(request.args.get("max_participants") or 250)
        except Exception:
            max_participants = 250

        # bornes de sécurité
        max_sessions = max(5, min(max_sessions, 200))
        max_participants = max(20, min(max_participants, 1000))

        magato = compute_magatomatique(
            flt,
            participant_q=participant_q,
            view=view,
            max_sessions=max_sessions,
            max_participants=max_participants,
        )

    participants = compute_participants_stats(flt)

    secteurs = []
    if getattr(current_user, "role", None) in ("finance", "financiere", "financière", "directrice", "admin_tech"):
        secteurs = [
            s[0]
            for s in (
                AtelierActivite.query.with_entities(AtelierActivite.secteur)
                .filter(AtelierActivite.is_deleted.is_(False))
                .distinct()
                .order_by(AtelierActivite.secteur.asc())
                .all()
            )
            if s and s[0]
        ]

    q = AtelierActivite.query.filter(AtelierActivite.is_deleted.is_(False))
    if flt.secteur:
        q = q.filter(AtelierActivite.secteur == flt.secteur)
    ateliers = q.order_by(AtelierActivite.secteur.asc(), AtelierActivite.nom.asc()).all()

    quartiers = Quartier.query.order_by(Quartier.nom.asc()).all()

    # Années disponibles (pour presets "année") dans le périmètre accessible
    try:
        role = getattr(current_user, "role", None)
        eff_secteur = flt.secteur
        if role == "responsable_secteur":
            eff_secteur = (getattr(current_user, "secteur_assigne", None) or "").strip() or eff_secteur

        year_expr = func.extract("year", func.coalesce(SessionActivite.rdv_date, SessionActivite.date_session))
        years_q = (
            db.session.query(year_expr.label("y"))
            .select_from(SessionActivite)
            .join(AtelierActivite, SessionActivite.atelier_id == AtelierActivite.id)
            .filter(AtelierActivite.is_deleted.is_(False))
        )
        if eff_secteur:
            years_q = years_q.filter(AtelierActivite.secteur == eff_secteur)
        years = [int(r.y) for r in years_q.distinct().order_by(year_expr.desc()).all() if r and r.y]
    except Exception:
        years = []

    return render_template(
        ["statsimpact/dashboard.html", "statsimpact_dashboard.html"],
        flt=flt,
        stats=stats,
        freq=freq,
        trans=trans,
        demo=demo,
        secteurs=secteurs,
        ateliers=ateliers,
        occupancy=occupancy,
        participants=participants,
        magato=magato,
        quartiers=quartiers,
        available_years=years,
    )



@bp.route("/stats-impact/magatomatique.xlsx", methods=["GET"])
@login_required
def magatomatique_export():
    if not _can_view():
        abort(403)

    flt = normalize_filters(dict(request.args), user=current_user)

    export_mode = (request.args.get("export_mode") or "flat").strip().lower()
    participant_q = (request.args.get("participant_q") or "").strip() or None
    view = (request.args.get("magato_view") or "macro").strip().lower()
    try:
        max_sessions = int(request.args.get("max_sessions") or 40)
    except Exception:
        max_sessions = 40
    try:
        max_participants = int(request.args.get("max_participants") or 250)
    except Exception:
        max_participants = 250

    # bornes (export raisonnable)
    max_sessions = max(5, min(max_sessions, 400))
    max_participants = max(20, min(max_participants, 5000))

    # Mode "per_atelier" : export annuel 1 feuille = 1 atelier
    if export_mode in ("per_atelier", "per-atelier", "atelier"):
        wb = _build_magato_per_atelier_workbook(flt)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = "magatomatique_par_atelier.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    magato = compute_magatomatique(
        flt,
        participant_q=participant_q,
        view=view,
        max_sessions=max_sessions,
        max_participants=max_participants,
    )

    if magato.get("restricted"):
        abort(403)

    wb = Workbook()
    ws = wb.active
    ws.title = "Synthese"

    # En-têtes synthèse macro (secteurs)
    ws.append(["Synthèse par secteur"])
    ws.append(["Secteur", "Nb sessions", "Nb présences", "Participants uniques"])
    for r in (magato.get("macro") or {}).get("by_secteur", []):
        ws.append([r["secteur"], r["nb_sessions"], r["nb_presences"], r["nb_participants_uniques"]])

    ws.append([])
    ws.append(["Synthèse par atelier"])
    ws.append(["Secteur", "Atelier", "Nb sessions", "Nb présences", "Participants uniques"])
    for r in (magato.get("macro") or {}).get("by_atelier", []):
        ws.append([r["secteur"], r["atelier_nom"], r["nb_sessions"], r["nb_presences"], r["nb_participants_uniques"]])

    # Feuille participants (si dispo)
    if magato.get("participants"):
        ws2 = wb.create_sheet("Participants")
        ws2.append(["Participants (dans le périmètre filtré)"])
        ws2.append(["Nom", "Prénom", "Ville", "Quartier", "Nb présences", "1ère venue", "Dernière venue"])
        for p in magato["participants"]:
            fd = p.get("first_date")
            ld = p.get("last_date")
            ws2.append([
                p.get("nom",""),
                p.get("prenom",""),
                p.get("ville") or "",
                p.get("quartier") or "",
                int(p.get("nb_presences",0)),
                fd.strftime("%Y-%m-%d") if fd else "",
                ld.strftime("%Y-%m-%d") if ld else "",
            ])

    # Feuille matrice (si view=matrix)
    if magato.get("view") == "matrix" and magato.get("sessions") and magato.get("participants"):
        ws3 = wb.create_sheet("Matrice")
        sessions = magato["sessions"]
        participants = magato["participants"]
        matrix = magato.get("matrix") or {}

        header = ["Nom", "Prénom"] + [s["label"] for s in sessions]
        ws3.append(header)

        for p in participants:
            row = [p.get("nom",""), p.get("prenom","")]
            pid = int(p["id"])
            for s in sessions:
                sid = int(s["id"])
                row.append("1" if matrix.get((pid, sid)) else "")
            ws3.append(row)

        # Ajuste largeur colonnes
        for col_idx in range(1, len(header) + 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 16 if col_idx <= 2 else 12

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "magatomatique.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
